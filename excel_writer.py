"""
每日运行结束后，将结构化数据追加到本地 Excel 文件的新一行。
文件不存在时自动创建并写入表头；已存在时直接追加，不覆盖历史数据。
"""
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from config import DATA_DIR

EXCEL_PATH = os.path.join(DATA_DIR, "sentiment_report.xlsx")

COLUMNS = [
    ("日期",       18),
    ("分组",       14),
    ("监控范围",   28),
    ("积极%",       8),
    ("消极%",       8),
    ("中立%",       8),
    ("情绪概述",   50),
    ("热门话题",   50),
    ("突发事件",   40),
    ("异常情况",   40),
    ("总订阅数",   12),
    ("当前在线",   12),
    ("帖子数",     10),
    ("平均分值",   10),
    ("总评论数",   12),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_WRAP        = Alignment(wrap_text=True, vertical="top")


def _create_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "舆情日报"
    ws.freeze_panes = "A2"

    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = _HEADER_FONT
        cell.fill   = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    return wb


def append_row(
    date_str: str,
    group_name: str,
    subreddits: list[str],
    sentiment: dict,
    today_sub: dict,
    today_post: dict,
) -> None:
    os.makedirs(DATA_DIR, exist_ok=True)

    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = _create_workbook()
        ws = wb.active

    row = [
        date_str,
        group_name,
        " + ".join(f"r/{s}" for s in subreddits),
        sentiment.get("positive_pct", ""),
        sentiment.get("negative_pct", ""),
        sentiment.get("neutral_pct", ""),
        sentiment.get("sentiment_explanation", ""),
        sentiment.get("hot_topics", ""),
        sentiment.get("emergent_events", ""),
        sentiment.get("anomalies", ""),
        today_sub.get("total_subscribers", 0),
        today_sub.get("total_active_users", 0),
        today_post.get("post_count", 0),
        round(today_post.get("avg_score", 0), 1),
        today_post.get("total_comments", 0),
    ]

    ws.append(row)

    # 对新写入的行应用样式
    new_row = ws.max_row
    for col_idx in range(1, len(COLUMNS) + 1):
        ws.cell(row=new_row, column=col_idx).alignment = _WRAP

    wb.save(EXCEL_PATH)
    print(f"Excel 已更新：{EXCEL_PATH}（第 {new_row} 行）")
